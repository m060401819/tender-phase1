from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import app.models  # noqa: F401
from app.db.base import Base
from app.db.session import get_db
from app.main import app
from app.models import CrawlError, CrawlJob, NoticeVersion, SourceSite, TenderNotice


def _seed_data(session_factory: sessionmaker) -> None:
    now = datetime.now(timezone.utc)
    with session_factory() as session:
        source_1 = SourceSite(
            id=1,
            code="anhui_ggzy_zfcg",
            name="安徽省公共资源交易监管网（政府采购）",
            base_url="https://ggzy.ah.gov.cn/",
            description="report test source 1",
            is_active=True,
            supports_js_render=False,
            crawl_interval_minutes=1440,
            default_max_pages=2,
            schedule_enabled=True,
            schedule_days=1,
        )
        source_2 = SourceSite(
            id=2,
            code="example_source",
            name="Example Source",
            base_url="https://example.com/",
            description="report test source 2",
            is_active=False,
            supports_js_render=True,
            crawl_interval_minutes=10080,
            default_max_pages=1,
            schedule_enabled=False,
            schedule_days=7,
        )
        session.add_all([source_1, source_2])

        job_succeeded = CrawlJob(
            id=101,
            source_site_id=1,
            job_type="manual",
            status="succeeded",
            triggered_by="test",
            started_at=now - timedelta(hours=3),
            finished_at=now - timedelta(hours=2, minutes=50),
            pages_fetched=3,
            documents_saved=3,
            notices_upserted=2,
            deduplicated_count=0,
            error_count=0,
            message="ok",
            created_at=now - timedelta(hours=3),
            updated_at=now - timedelta(hours=3),
        )
        job_failed = CrawlJob(
            id=102,
            source_site_id=1,
            job_type="manual",
            status="failed",
            triggered_by="test",
            started_at=now - timedelta(hours=2),
            finished_at=now - timedelta(hours=1, minutes=50),
            pages_fetched=1,
            documents_saved=1,
            notices_upserted=0,
            deduplicated_count=0,
            error_count=1,
            message="timeout",
            created_at=now - timedelta(hours=2),
            updated_at=now - timedelta(hours=2),
        )
        job_retry = CrawlJob(
            id=103,
            source_site_id=1,
            job_type="manual_retry",
            status="succeeded",
            triggered_by="admin-retry",
            retry_of_job_id=102,
            started_at=now - timedelta(hours=1),
            finished_at=now - timedelta(minutes=55),
            pages_fetched=1,
            documents_saved=1,
            notices_upserted=1,
            deduplicated_count=0,
            error_count=0,
            message="retry ok",
            created_at=now - timedelta(hours=1),
            updated_at=now - timedelta(hours=1),
        )
        job_old = CrawlJob(
            id=104,
            source_site_id=2,
            job_type="manual",
            status="failed",
            triggered_by="test",
            started_at=now - timedelta(hours=30),
            finished_at=now - timedelta(hours=29, minutes=50),
            pages_fetched=1,
            documents_saved=1,
            notices_upserted=0,
            deduplicated_count=0,
            error_count=1,
            message="old failed",
            created_at=now - timedelta(hours=30),
            updated_at=now - timedelta(hours=30),
        )
        session.add_all([job_succeeded, job_failed, job_retry, job_old])

        notice = TenderNotice(
            id=201,
            source_site_id=1,
            external_id="RPT-001",
            project_code="RPT-001",
            dedup_hash="rpt-001",
            title="report notice",
            notice_type="announcement",
            issuer="issuer",
            region="合肥",
            published_at=now - timedelta(hours=2),
            deadline_at=now + timedelta(days=5),
            budget_amount=None,
            budget_currency="CNY",
            summary=None,
            first_published_at=now - timedelta(hours=2),
            latest_published_at=now - timedelta(hours=2),
            current_version_id=301,
            created_at=now - timedelta(hours=2),
            updated_at=now - timedelta(hours=2),
        )
        version = NoticeVersion(
            id=301,
            notice_id=201,
            raw_document_id=None,
            version_no=1,
            is_current=True,
            content_hash="rpt-hash-1",
            title="report notice",
            notice_type="announcement",
            issuer="issuer",
            region="合肥",
            published_at=now - timedelta(hours=2),
            deadline_at=now + timedelta(days=5),
            budget_amount=None,
            budget_currency="CNY",
            structured_data={},
            change_summary=None,
            created_at=now - timedelta(hours=1),
            updated_at=now - timedelta(hours=1),
        )
        session.add_all([notice, version])

        latest_error = CrawlError(
            id=401,
            source_site_id=1,
            crawl_job_id=102,
            raw_document_id=None,
            stage="fetch",
            url="https://ggzy.ah.gov.cn/list",
            error_type="FetchTimeout",
            error_message="latest timeout",
            traceback=None,
            retryable=True,
            occurred_at=now - timedelta(hours=1, minutes=55),
            resolved=False,
            created_at=now - timedelta(hours=1, minutes=55),
            updated_at=now - timedelta(hours=1, minutes=55),
        )
        session.add(latest_error)
        session.commit()


def _build_client(tmp_path: Path) -> tuple[TestClient, object]:
    db_url = f"sqlite+pysqlite:///{tmp_path / 'reports_api.db'}"
    engine = create_engine(db_url)
    Base.metadata.create_all(engine)

    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False)
    _seed_data(session_factory)

    def override_get_db():
        db = session_factory()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)
    return client, engine


def test_source_ops_report_xlsx_returns_expected_fields_and_rows(tmp_path: Path) -> None:
    client, engine = _build_client(tmp_path)
    try:
        response = client.get("/reports/source-ops.xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "source-ops-report-" in response.headers.get("content-disposition", "")

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0] == (
            "source_code",
            "source_name",
            "official_url",
            "is_active",
            "schedule_enabled",
            "schedule_days",
            "today_crawl_job_count",
            "today_success_count",
            "today_failed_count",
            "today_partial_count",
            "today_new_notice_count",
            "last_job_status",
            "last_job_finished_at",
            "last_error_message",
            "last_retry_status",
        )
        row_map = {row[0]: row for row in rows[1:]}
        source_1 = row_map["anhui_ggzy_zfcg"]
        assert source_1[6] == 3
        assert source_1[7] == 2
        assert source_1[8] == 1
        assert source_1[9] == 0
        assert source_1[10] == 1
        assert source_1[11] == "succeeded"
        assert source_1[13] == "latest timeout"
        assert source_1[14] == "succeeded"
    finally:
        app.dependency_overrides.clear()
        engine.dispose()


def test_source_ops_report_xlsx_supports_source_code_filter(tmp_path: Path) -> None:
    client, engine = _build_client(tmp_path)
    try:
        response = client.get("/reports/source-ops.xlsx", params={"source_code": "example_source"})
        assert response.status_code == 200
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][0] == "example_source"
    finally:
        app.dependency_overrides.clear()
        engine.dispose()
